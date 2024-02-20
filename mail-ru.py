import os # Для работы с файлами.
import re # Работа с поиском по регулярным выражениям.
import time  # Для задержки выполнения скрипта.
import xlrd # Работа устаревшими файлами Excel.
import magic  # Определение MIME-типа файлов на основе их содержимого.
import pickle  # Для работы с cookies.
import asyncio # Для асинхронной работы телеграмма.
import random  # Для генерации рандомного юзерагента.
import zipfile # Для работы с архивами.
import requests # Для работы с запросами.
import openpyxl # Работа с Excel.
import textract  # Извлечение текста из различных документов.
import mimetypes  # Определение MIME-типа файла на основе его расширения.
import pdfplumber # Работа с PDF.
import pytesseract  # Распознавание текста на изображениях с использованием Tesseract OCR.
from PIL import Image  # Работы с изображениями (Python Imaging Library).
from telegram import Bot # Для работы с телеграмма.
from bs4 import BeautifulSoup # Парсинг HTML
from datetime import datetime  # Для работы с временем
from fake_useragent import UserAgent  # Для генерации юзерагента.
from urllib.parse import urlparse, unquote # Работа с URL.
from selenium.webdriver.common.action_chains import ActionChains  # Класс для выполнения сложных действий веб-драйвера, таких как наведение и клик.
from selenium.common.exceptions import NoSuchElementException, TimeoutException  # Исключения, возникающие при отсутствии элемента или превышении времени ожидания.
from selenium.webdriver.support import expected_conditions as EC  # Модуль с ожиданиями (предусловиями), используемыми вместе с WebDriverWait.
from selenium.webdriver.support.ui import WebDriverWait  # Класс для ожидания определенных условий веб-страницы перед выполнением действий.
from selenium.webdriver.common.keys import Keys  # Класс для взаимодействия с клавишами на клавиатуре.
from selenium.webdriver.common.by import By  # Класс для использования различных стратегий поиска элементов.
from selenium.webdriver.chrome.options import Options  # Класс для управления опциями веб-драйвера Chrome.
from selenium import webdriver  # Основной класс для веб-драйвера, который предоставляет методы для взаимодействия с веб-страницами.



useragent = UserAgent()
chrome_options = Options() # Опции для Chrome
# Добавление опции --no-sandbox (используется в среде, где отсутствует поддержка среды исполнения)
chrome_options.add_argument('--no-sandbox')
chrome_options.add_experimental_option("detach", True)
# Включение режима "фонового запуска" (headless mode)
#chrome_options.add_argument('--headless')
# Добавление опции для сохранения профиля браузера в определенной директории
chrome_options.add_argument("user-data-dir=/srv/Selenium_python/browser_settings/")
# Другие опции можно добавить по аналогии, например, отключение уведомлений
chrome_options.add_argument("--disable-notifications")
desired_user_agent = "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
chrome_options.add_argument(f'user-agent={desired_user_agent}') # Добавление юзерагента.
print(f"Установлен User-Agent: {desired_user_agent}") # Вывод в консоль User-Agent
chrome_options.add_argument("--disable-blink-features=AutomationControlled") # Отключение режима автоматизации
# chrome_options.add_argument('--proxy-server=67.43.228.253:1035') # Добавление опции --proxy-server
driver = webdriver.Chrome(options=chrome_options)# Инициализация драйвера Chrome с использованием опций.
driver.maximize_window() # Максимизация окна браузера
# URL-адрес, который мы хотим открыть
url = "https://account.mail.ru/login" # Авторизация
mail_url = "https://e.mail.ru/inbox" # С этой страницы программа парсит почту
login = os.environ.get("login")  # Логин
# Получаем зашифрованный пароль из переменной окружения
password = os.environ.get("password") # Пароль
# Функция для отправки сообщения в Telegram
token = os.environ.get("token")
# Указываем ID чата, куда отправлять сообщение (это может быть ваш ID или ID группового чата)
chat_id = "368228657" 
downloads_folder = "/root/Downloads/" # Папка с загруженными файлами
file_path = "/root/Downloads/output.txt"  # Путь и имя файла по умолчанию



def main():
    try:
        driver.get(mail_url)
        print(f"{current_time()} Открыта страница: {mail_url}.")
        time.sleep(2)  # Подождать, чтобы страница снова успела загрузиться
        # Функция для фильтрации по вложениям и непрочитанным.
        filter(driver)
        # Функция для получения количества писем. После получения списка вызывает следующую функцию для парсинга писем со вложениями.
        get_email_count(driver)
        time.sleep(2)
        print(f"{current_time()} Письмо обработано. Новый запрос.")
        main()
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке main(): {e}.")



def current_time(): # Функция для получения текущего времени
    update_time = time.strftime("%H:%M:%S.", time.localtime())
    return update_time



async def send_message_to_user(token, chat_id, current_url, text): # Функция для отправки сообщения в Telegram
    try:
        # Создаем экземпляр бота с использованием токена
        bot = Bot(token=token)
        # Отправляем сообщение пользователю
        message = (f"Время в системе: {current_time()}\n{text}\n\n<a href=\"{current_url}\">Ссылка на письмо</a>.")
        await bot.send_message(chat_id=chat_id, text=message, parse_mode='HTML')
        print(f"{current_time()} Сообщение об ошибке успешно отправлено в Telegram.")
    except Exception as e:
        print(f"{current_time()} Ошибка при отправке сообщения в Telegram: {e}")



def add_russian_language(driver): # Перевод браузера на русский.
    try:
        driver.get(url); print(f"{current_time()} Открытие веб-страницы {url}") # Открытие веб-страницы
        time.sleep(2)
        # Поиск текста на странице.
        found_element = driver.find_elements(By.XPATH, f"//*[contains(text(), 'Войти в аккаунт')]")
        count_elements = 0 
        if found_element:
            print(current_time(), "Элемент 'Войти в аккаунт' найден.")
            count_elements = 1
        else:
            print(current_time(), "Элемент 'Войти в аккаунт' не найден.")
        found_element2 = driver.find_elements(By.XPATH, f"//*[contains(text(), 'Добавить аккаунт')]")
        if found_element2:
            print(current_time(), "Элемент 'Добавить аккаунт' найден. Значит мы авторизованы.")
            count_elements = 1
        else:
            print(current_time(), "Элемент 'Добавить аккаунт' не найден.")
        time.sleep(2)
        if count_elements == 0: # Если элементы не найдены переводим браузер на русский язык
            print(current_time(), "Браузер не переведен на русский язык.")
            time.sleep(1)
            # Открываем страницу настроек языков
            driver.get('chrome://settings/languages')
            time.sleep(2)
            # Переключаемся на новую вкладку (последнюю в списке)
            driver.switch_to.window(driver.window_handles[-1])
            # Добавляем русский язык.
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys("russian").perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform(); time.sleep(0.3)
            webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform(); time.sleep(0.3)
            # Открываем новую вкладку с помощью JavaScript
            driver.execute_script("window.open('', '_blank');")
            # Переключаемся на новую вкладку (последнюю в списке)
            driver.switch_to.window(driver.window_handles[-1])
            # Закрываем фоновую вкладку (первую в списке)
            driver.switch_to.window(driver.window_handles[0])
            driver.close()
            # Переключаемся обратно на новую вкладку
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(1)
        else:
            print(current_time(),"Русский язык Активен. Передано управление в main")
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке add_russian_language(): {e}.")



def authorization(driver): # Функция для авторизации
    try:
        driver.get(url); print(f"{current_time()} Открытие веб-страницы {url}") # Открытие веб-страницы
        time.sleep(2)
        # Поиск текста на странице.
        found_element = driver.find_elements(By.XPATH, f"//*[contains(text(), 'Войти в аккаунт')]")
        found_element2 = driver.find_elements(By.XPATH, f"//*[contains(text(), 'Добавить аккаунт')]")
        if found_element:
            print(current_time(), "Элемент 'Войти в аккаунт' найден.")
            # Ввод логина
            for i in login:
                webdriver.ActionChains(driver).send_keys(i).perform()
                time.sleep(random.uniform(0.3, 0.7))
            print(current_time(), "Логин введен.")
            # Клик по элементу "Ввести пароль".
            driver.find_element(By.XPATH, '//span[contains(text(), "Ввести пароль")]').click()
            time.sleep(2)
            # Ввод пароля
            for i in password: # Пароль для входа
                webdriver.ActionChains(driver).send_keys(i).perform()
                time.sleep(random.uniform(0.3, 0.7))
            print(current_time(),"Пароль введен.")
            # Клик по элементу "Войти".
            driver.find_element(By.XPATH, '//span[contains(text(), "Войти")]').click()
            time.sleep(1)
            
        elif found_element2:
            if found_element2:
                print(current_time(), "Элемент 'Добавить аккаунт' найден. Переходим в почту.")
                driver.get(mail_url)
                time.sleep(1)
        else:
            print(current_time(),"Аккаунт авторизован. Передано управление в main.")
        time.sleep(5)
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке authorization(): {e}.")



def mark_an_error(driver): # В письме присутствует ошибка. Например отсутствует ссылка на купец.
    try:
        error_element = driver.find_element(By.XPATH, '//*[@id="app-canvas"]/div/div[1]/div[1]/div/div[2]/span/div[2]/div/div/div/div/div[1]/div/div[2]/div[5]/div/div/div/span/span[2]')
        error_element.click()
        print(f"{current_time()} Троеточие успешно нажато.")
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[contains(text(), 'Пометить флажком')]").click()
        print(f"{current_time()} Пометить флагом успешно нажато.")
        error_element.click()
        print(f"{current_time()} Повторное нажатие на троеточие.")
        driver.find_element(By.XPATH, "//span[contains(text(), 'Пометить непрочитанным')]").click()
        print(f"{current_time()} Письмо помечено непрочитанным.")
    except NoSuchElementException:
        print(f"{current_time()} Элемент не найден. Обновляем страницу и повторяем попытку.")
        driver.refresh()
        time.sleep(2)
        mark_an_error(driver)
    except Exception as e:
        print(f"{current_time()} except mark_an_error(): {e}.")



def filter(driver): # Функция для фильтрации по вложениям и непрочитанным.
    try:
        element_filter = driver.find_element(By.CLASS_NAME, 'filters-control__filter-text')
        if element_filter.is_displayed(): # Нажатие на фильтровать по вложениям.
            element_filter.click(); time.sleep(1)
            elements_attachments = driver.find_elements(By.CLASS_NAME, 'list-item__text')
            for element in elements_attachments:
                if element.text.strip() == 'С вложениями':
                    element.click()
                    print(current_time(), "Письма отфильтрованы по вложениям.")
                    break
            else:
                print(current_time(), "Элемент 'С вложениями' не найден или не отображен.")
        if element_filter.is_displayed(): # Нажатие на фильтровать по непрочитанным.
            element_filter.click(); time.sleep(1)
            # Находим элемент с текстом 'Непрочитанные' по XPath
            xpath_unread = "//span[@class='list-item__text' and text()='Непрочитанные']"
            element_unread = driver.find_element(By.XPATH, xpath_unread)
            if element_unread.is_displayed():
                element_unread.click(); time.sleep(1)
                print(current_time(), "Письма отфильтрованы по непрочитанным.")
            else:
                print(current_time(), "Элемент 'Непрочитанные' не найден или не отображен.")
        else:
            print(current_time(), "Элемент 'Фильтр' не найден или не отображен.")
    except Exception as e:
        print(f"Ошибка в блоке filter(): {e}.")



def get_email_count(driver): # Функция для получения количества писем. Возвращает список непрочитанных писем со вложениями.
    try:
        elements_flag = [] # Сборщик непрочитанных писем со вложениями и флагом. По умолчанию пустой список.
        # Все непрочитанные письма на странице.
        elements = driver.find_elements(By.CLASS_NAME, 'llc') 
        # Все непрочитанные письма на странице.
        elements_with_title = [el for el in elements if "Пометить прочитанным" in el.get_attribute("outerHTML")]
        # Все непрочитанные письма со вложениями на странице.
        elements_investment = [el for el in elements_with_title if "С вложениями" in el.get_attribute("outerHTML")] 
        # Все непрочитанные письма со вложениями и флагом на странице.
        elements_flag = [el for el in elements_with_title if "Снять флажок" in el.get_attribute("outerHTML")]
        if elements_investment and elements_flag:
            # Удалить элементы с флагом из списка с вложениями. Бот будет игнорировать все письма помеченные флагом.
            elements_investment = [el for el in elements_investment if el not in elements_flag]
            print(f"{current_time()} Обнаружены письма с флагом. Такие письма будут игнорироваться.")
        elements_investment_result = elements_investment
        print(f"{current_time()} Количество писем: {len(elements)}. Из них непрочитанных: {len(elements_with_title)}.",
                f"Имеют вложения: {len(elements_investment)}. Проигнорировано писем с флагом: {len(elements_flag)}.")
        if len(elements_investment) == 0: # Если непрочитанных и не помеченных писем нет.
            if len(elements_with_title) == len(elements_flag): # И количество писем со вложениями и письма с флагом равны. Проматываем вниз 5 раз.
                element = driver.find_element(By.CSS_SELECTOR, "._13g9TwPADGlVYjejUp8ylL._1ejZpbKssklxkVV-RWwRM8")
                element.click()
                # Количество прокруток
                num_scrolls = 5
                # Цикл для выполнения прокруток
                for _ in range(num_scrolls):
                    webdriver.ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform(); time.sleep(1)
                    # Повторно получаем список всех писем
                    elements = driver.find_elements(By.CLASS_NAME, 'llc') 
                    # Все непрочитанные письма на странице.
                    elements_with_title = [el for el in elements if "Пометить прочитанным" in el.get_attribute("outerHTML")]
                    # Все непрочитанные письма со вложениями на странице.
                    elements_investment = [el for el in elements_with_title if "С вложениями" in el.get_attribute("outerHTML")] 
                    # Все непрочитанные письма со вложениями и флагом на странице.
                    elements_flag = [el for el in elements_with_title if "Снять флажок" in el.get_attribute("outerHTML")]
                    if elements_investment and elements_flag:
                        # Удалить элементы с флагом из списка с вложениями. Бот будет игнорировать все письма помеченные флагом.
                        elements_investment = [el for el in elements_investment if el not in elements_flag]
                        print(f"{current_time()} Обнаружены письма с флагом. Такие письма будут игнорироваться.")
                    elements_investment_result = elements_investment
                    print(f"{current_time()} Непрочитанных: {len(elements_with_title)}. Проигнорировано писем с флагом: {len(elements_flag)}. Спускаемся по странице вниз в поисках новых писем.")
                    # Проверяем, есть ли новые непрочитанные письма без флага
                    if len(elements_investment_result) > 0:
                        return open_last_investment_email(driver, elements_investment_result)
                print(f"{current_time()} Новых писем без флагов не найдено. Ожидание минута.")
                time.sleep(60)
                main() # Проматали страницу 5 раз вниз. Подождём минуту и вернёмся к новому кругу.
        return open_last_investment_email(driver, elements_investment_result)
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке get_email_count(): {str(e)}.")



def link_found(driver): # Функция ищет ссылку на купец на странице. Использует разные методы поиска. Каждый из них срабатывал на разных письмах.
    try:
        methods = [
            (By.ID, "mcntmcntmcntmcntmcntacceptLink_mr_css_attr"),
            (By.XPATH, "//div[contains(text(), 'По правилам нашей компании, мы выбираем лучшего поставщика, исходя из сравнения предложений по этой рассылке.')]/a"),
            (By.CSS_SELECTOR, "div.letter-blockquote__body blockquote div:nth-child(5) a"),
            (By.ID, "acceptLink_mr_css_attr"), 
            (By.XPATH, "//span[contains(text(), 'По правилам нашей компании, мы выбираем лучшего поставщика, исходя из сравнения предложений по этой рассылке.')]/a"),
            (By.XPATH, "//p[@class='MsoNormal_mr_css_attr' and contains(text(), 'По правилам нашей компании, мы выбираем лучшего поставщика, исходя из сравнения предложений по этой рассылке.')]//a"),
        ]
        for attempt in range(1, 4):  # Цикл из 3 попыток
            for method in methods:
                try:
                    element = driver.find_element(By.CLASS_NAME, "letter__header-details")
                    element.click()
                    webdriver.ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform(); time.sleep(0.5)
                    element_before = driver.find_element(*method)
                    actions = ActionChains(driver)
                    actions.move_to_element(element_before).perform()
                    time.sleep(0.5)
                    link = element_before.get_attribute('href')
                    print(f"{current_time()} Ссылка на купца найдена.")
                    return link
                except NoSuchElementException:
                    pass
            # Если не найдено, обновляем страницу и ждем 2 секунды перед повторной попыткой
            print(f"{current_time()} Ссылка на купца не найдена, обновим страницу. Попытка {attempt} из 3.")
            driver.refresh()
            time.sleep(2)
        print(f"{current_time()} Ссылки на купца нет. Письмо будет помечено.")
        # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
        text = "В письме нет ссылки на купца."
        current_url = driver.current_url
        asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
        mark_an_error(driver)
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке link_found(): {str(e)}.")



def open_last_investment_email(driver, elements_investment): # Эта функция выдёргивает последнее письмо и распознаёт его.
    try:
        print(f"{current_time()} Принято на обработку {len(elements_investment)} писем.")
        # Удаление файлов из папки загрузок
        for file_name in os.listdir(downloads_folder):
            file_path = os.path.join(downloads_folder, file_name)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                    print(f"{current_time()} Удален файл: {file_name}")
            except Exception as e:
                print(f"{current_time()} Ошибка В блоке open_last_investment_email(). Ошибка при удалении файла {file_path}: {e}.")
        # Получаем последний элемент списка
        last_element = elements_investment[-1]
        print(f"{current_time()} Обрабатывается письмо: {last_element.text.splitlines()[0]}.")
        time.sleep(3)
        last_element.click()  # Кликнуть на ссылку последнего письма
        time.sleep(3)  # Подождать, чтобы страница успела загрузиться
        download_link = link_search(driver) # Узнаем ссылку на скачку вложений.
        letter_body_text = driver.find_element(By.CLASS_NAME, "letter-body__body-wrapper").text
        text_body = process_letter_body_text(letter_body_text)
        # Находим контактные данные почты.
        span_element = driver.find_element(By.CLASS_NAME, 'letter-contact')
        # Получаем данные из атрибута title и текста внутри тега
        email = span_element.get_attribute('title')
        link = link_found(driver)  # Ищем ссылку на купца.
        if link:
            # Если ссылка найдена перейдем на купец, и запишем содержимое.
            kup_get(driver, link)
            write_to_file(email=email, Содержимое_письма=text_body)
            open_link_in_browser(driver, download_link)
            found_letter_is_correct(driver)
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке open_last_investment_email(): {str(e)}.")



def link_search(driver, max_attempts=3): # Данная функция ищет ссылку на вложение со странице почты. Так как на странице всегда есть ссылка на файл, в случае если функция её не нашла, запишем на какой мы странице и остановим всю программу для отладки функции.
    for attempt in range(max_attempts):
        try:
            a_element = driver.find_element(By.CLASS_NAME, 'attach-list__controls-element-download') 
            download_link = a_element.get_attribute('href')
            print(f"{current_time()} Ссылка на файл найдена.")
            return download_link
        except NoSuchElementException as e:
            print(f"{current_time()} Ссылка на файл не найдена, обновляем страницу. Попытка {attempt + 1} из {max_attempts}.")
            driver.refresh(); time.sleep(2)  # Дополнительное ожидание перед новой попыткой поиска элемента
        except Exception as e:
            current_url = "Ссылка на страницу с ошибкой - "+ driver.current_url  # Инициализируем переменную current_url в ней создадим ссылку на страницу.
            error_message = f"{current_time()} Ошибка в блоке link_search() на странице {current_url}: {str(e)}."
            write_to_file(link_search=current_url)
            raise SystemExit("Программа остановлена из-за ошибки в блоке link_search.")



def process_letter_body_text(letter_body_text): # Текст извлекаемый из письма сложно читаемый. Чистим с помощью регулярных выражений.
    try:
        all_text = re.sub(r'(?<![\.,:;!?])$', '.', letter_body_text, flags=re.MULTILINE) # Добавляем точку в конце
        all_text = re.sub(r'\. (\w)', lambda x: '. ' + x.group(1).capitalize(), all_text) # Вставляем точку перед заглавной буквой
        all_text = re.sub(r'\n\s*\n', '\n', all_text) # Удаляем лишние переносы строк
        all_text = re.sub(r'</\.', '', all_text) # Удаляем лишние символы
        all_text = re.sub(r'\.{2,}', '.', all_text) # Удаляем лишние точки
        all_text = re.sub(r'\.+', '.', all_text) # Удаляем лишние точки
        all_text = re.sub(r'\.\s\.', '.', all_text) # Удаляем лишние точки
        all_text = re.sub(r'--\.', '', all_text) # Удаляем лишние точки
        all_text = re.sub(r'  ', '', all_text) # Удаляем лишние пробелы
        all_text = re.sub(r'\. \)', ')', all_text) # Удаляем лишние символы
        all_text = re.sub(r'\s+\.', '', all_text) # Удалит точку и пробел, если перед ней есть пробел.
        # Возвращаем результат
        return all_text
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке process_letter_body_text(). Произошла ошибка при обработке текста: {e}.")



def open_link_in_browser(driver, download_link): # Скачивает файл по ссылке. Функции приходит ссылка на все файлы, если их несколько то это архив, далее распаковка и переименовывание файлов. Отправляет их дальше на распознование. Файл output.txt игнорирует.
    try:
        # Открываем страницу по указанной ссылке
        driver.get(download_link)
        # Ожидаем, чтобы браузер успел скачать файл (может потребоваться настройка)
        time.sleep(5)
        # Проверяем новые файлы в папке загрузок, игнорируя файл output.txt
        new_files = [f for f in os.listdir(downloads_folder) if os.path.isfile(os.path.join(downloads_folder, f)) and f.lower() != "output.txt"]
        if new_files:
            file_name = new_files[0]
            # Заменяем пробелы на подчеркивания
            new_file_name = file_name.replace(" ", "_")
            # Формируем новый путь к файлу с учетом измененного имени
            new_file_path = os.path.join(downloads_folder, new_file_name)
            # Переименовываем файл
            os.rename(os.path.join(downloads_folder, file_name), new_file_path)
            print(f"{current_time()} Скачан новый файл: {new_file_name}")
            return process_downloaded_file(new_file_path)
        else:
            print(f"{current_time()} Новые файлы не найдены (или найден только output.txt).")
            return None
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке open_link_in_browser(). Ошибка при скачивании файла: {e}.")
        return None



def process_downloaded_file(file_path): # Анализирует расширение скачанного файла и выбирает соответствующий способ обработки. Если будет несколько файлов, будет вызвана несколько раз, пока не будут обработаны все файлы.
    try:
        if file_path:
            # Получаем расширение файла из пути
            file_extension = os.path.splitext(file_path)[-1].lower()
            # Вызываем соответствующую функцию в зависимости от расширения
            if file_extension == ".pdf":
                print(f"{current_time()} Получен PDF-файл.")
                process_pdf(file_path)
            elif file_extension in (".jpg", ".jpeg", ".png", ".gif"):
                print(f"{current_time()} Получено изображение.")
                process_image(file_path)
            elif file_extension in (".doc", ".docx"):
                print(f"{current_time()} Получен Word. ")
                process_document(file_path)
            elif file_extension in (".xlsx", ".xlsm", ".xlsb"): 
                print(f"{current_time()} Получен Excel.")
                read_excel(file_path)
            elif file_extension in (".xls"): 
                print(f"{current_time()} Получен Excel.")
                convert_to_xlsx(file_path)
            elif file_extension in (".zip"):
                print(f"{current_time()} Получен Архив.")
                unzip_archive(file_path)
            else:
                print(f"{current_time()} Неподдерживаемое расширение файла. ")
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке process_downloaded_file(). Ошибка при обработке скачанного файла: {e}.")
        return None



def process_files(file_paths): # Работает со списком файлов, исключая архив.
    try:
        # Создаем новый список для архивов
        archive_files = [file_path for file_path in file_paths if file_path.endswith('.zip')]
        results = None
        # Удаляем архивы из исходного списка
        file_paths = [file_path for file_path in file_paths if not file_path.endswith('.zip')]
        # Обрабатываем оставшиеся файлы (не архивы)
        for file_path in file_paths:
            # Вызываем функцию для обработки файла
            results = process_downloaded_file(file_path)
            # Делаем что-то с результатами, например, выводим их
            if results is not None:
                print(f"Результаты для файла {file_path}: {results}")
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке process_files(). Ошибка при обработке файлов: {e}.")



def unzip_archive(new_file_path): # Разбирает архив. Если там фото и документ, фото и архив будет удалены. Документ распознан.
    try:
        folder_path = os.path.dirname(new_file_path)
        with zipfile.ZipFile(new_file_path, 'r') as zip_ref:
            zip_ref.extractall(folder_path)
        rename_files(folder_path)
        extracted_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path)]
        document_files = [file for file in extracted_files if file.endswith((".xlsx", ".xlsm", ".xlsb", ".xls", ".doc", ".docx", ".jpg", ".jpeg", ".png", ".gif", ".pdf"))]
        if len(document_files) >= 2:
            os.remove(new_file_path)
            print(f"{current_time()} Архив {new_file_path} успешно удален.")
            return process_files(document_files)
        else:
            print(f"{current_time()} В архиве нет достаточного количества документов.")
            return process_files(document_files)
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке unzip_archive(). Ошибка при разархивации файла {new_file_path}: {str(e)}.")
        return None



def rename_files(folder_path): # Заменяет пробелы в названии файлов на _.
    try:
        # Получаем список всех файлов в папке
        files = os.listdir(folder_path)
        # Переименовываем каждый файл
        for file_name in files:
            old_path = os.path.join(folder_path, file_name)
            # Изменяем имя файла, убирая пробелы и вставляя символ подчеркивания
            new_name = file_name.replace(' ', '_')
            # Разделяем имя файла и его расширение
            base_name, extension = os.path.splitext(new_name)
            # Преобразуем расширение в нижний регистр
            extension = extension.lower()
            # Собираем новое имя файла с обновленным расширением
            new_name = f"{base_name}{extension}"
            new_path = os.path.join(folder_path, new_name)
            # Переименовываем файл
            os.rename(old_path, new_path)
    except Exception as e:
        print(f"Ошибка В блоке rename_files(): {str(e)}.")



def process_pdf(file_path): # Функция распознаёт текст на pdf.
    print(f"{current_time()} Обработка PDF: {file_path}")
    try:
        with pdfplumber.open(file_path) as pdf:
            text_content = ''  # Изменено имя переменной на text_content
            for page in pdf.pages:
                text_content += page.extract_text()
        text_content = re.sub(r'\n\s*\n', '\n', text_content)
        # Проверим что распознали.
        process_and_delete_file(Link=file_path, result=text_content)
        return
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке process_pdf(). Ошибка при извлечении текста из PDF: {e}.")
        return None



def process_image(image_path): # Распознаём картинку. Функция настроена на русский язык.
    try:
        print(f"{current_time()} Обработка изображения: {image_path}")
        # Открываем изображение
        with Image.open(image_path) as img:
            # Распознаем текст с использованием Tesseract
            recognized_text = pytesseract.image_to_string(img, lang='rus')
            recognized_text = re.sub(r'\n\s*\n', '\n', recognized_text) # Удалит все лишние переносы.
            # Проверим что распознали.
            process_and_delete_file(Link=image_path, result=recognized_text)
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке process_image(). Ошибка при обработке изображения: {e}.")



def process_document(file_path): # Функция распознаёт doc и docx.
    try:
        mime = magic.Magic()
        mime_type, encoding = mimetypes.guess_type(file_path)
        print(f"{current_time()} Обработка Word: {file_path}")
        if mime_type in ('application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'):
            try:
                content = textract.process(file_path).decode('utf-8')
                content = re.sub(r'\n\s*\n', '\n', content) # Удалит все лишние переносы.
                # Проверим что распознали.
                process_and_delete_file(Link=file_path, result=content)
            except Exception as e:
                print(f"{current_time()} Ошибка В блоке process_document(). Ошибка чтения документа: {e}.")
        else:
            print(f'{current_time()} Ошибка В блоке process_document(). Файл не является документом Word.')
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке process_document(). Ошибка при чтении Word-файла: {e}.")



def convert_to_xlsx(input_file): # Функция конвертирует .xls в .xlsx
    try:
        # Открываем .xls файл с помощью xlrd
        xls_workbook = xlrd.open_workbook(input_file)
        # Создаем новый .xlsx файл с помощью openpyxl
        xlsx_workbook = openpyxl.Workbook()
        # Создаем новый лист на первой позиции
        xlsx_sheet = xlsx_workbook.create_sheet(title="TDSheet", index=0)
        # Копируем данные из .xls в .xlsx
        for sheet_index in range(xls_workbook.nsheets):
            xls_sheet = xls_workbook.sheet_by_index(sheet_index)
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row, col)
                    xlsx_sheet.cell(row=row+1, column=col+1).value = cell_value
        # Определяем путь и имя нового .xlsx файла
        input_dir, input_filename = os.path.split(input_file)
        input_name, input_ext = os.path.splitext(input_filename)
        file_path = os.path.join(input_dir, f"{input_name}.xlsx")
        # Сохраняем .xlsx файл
        xlsx_workbook.save(file_path)
        print(f"{current_time()} Файл '.xls' успешно конвертирован в формат '.xlsx")
        # Удаляем старый .xls файл
        os.remove(input_file)
        print(f"{current_time()} Старый файл '.xls' удален. Распознаем '.xlsx")
        # Передаем ссылку на новый файл в функцию read_excel
        read_excel(file_path)
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке convert_to_xlsx(). Ошибка при конвертации и чтении в формате .xlsx: {e}.")



def read_excel(file_path): # Функция распознаёт Excel.
    try:
        print(f"{current_time()} Обработка Excel: {file_path}")
        # Открываем файл Excel
        workbook = openpyxl.load_workbook(file_path)
        # Выбираем первый лист в книге
        sheet = workbook.active
        # Извлекаем текст из каждой ячейки
        content = ""
        for row in sheet.iter_rows(values_only=True):
            content += " ".join(str(cell) for cell in row) + "\n"
        content = re.sub(r'\bNone\b', '', content) # Удалит все None из текста.
        content = re.sub(r'\n\s*\n', '\n', content) # Удалит все лишние переносы.
        # Проверим что распознали.
        process_and_delete_file(Link=file_path, result=content)                    
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке read_excel(). Ошибка при чтении Excel-файла: {e}.")



def process_and_delete_file(Link, result): # Фильтрует чтобы счёт был выставлен на ту компанию на которую осуществляется закупка. А так же какой тип документа и НДС. Если это не счёт и не кп документ будет удалён.
    try:
        # Получаем название компании из файла
        company = read_from_file('Компания')
        print(f"{current_time()} Компания: {company}.")  
        company_matches = re.findall(r'["\']([^"\']*)["\']', company)  # Находим текст внутри кавычек.
        # Если список company_matches не пустой, берем первый элемент, иначе используем исходную строку.
        company = company_matches[0] if company_matches else company
        # Удаляем указанные слова и приводим к нижнему регистру.
        company = re.sub(r'\b(ИП|индивидуальный предприниматель|ООО|общество с ограниченной ответственностью|АО|акционерное общество)\b', '', company, flags=re.IGNORECASE)
        company = company.strip().lower()    
        # Поиск слов по словарю
        payment_invoice = ["счёт ", "счет ", "счёт на оплату", "счет на оплату"]
        commercial_proposal = ["кп ", "коммерческое предложение"]
        # Удаляет пробелы, которые находятся перед цифрами, за которыми следует запятая. 
        result = re.sub(r'(?<=\d)\s+(?=\d{3},\d)', '', result) # Удаляет пробелы между тысячами.
        result = re.sub(r'\b(?<!,)(\d+)\s+(?=\d+,\d{2}\b)', r'\1', result) # Удаляет пробелы в миллионах.
        # Уберём 20% НДС а так же сумму после, для удобства нейросети.
        result = re.sub(r'\b\d+(?:\.\d+)?%(\s*\d+(?:,\d+)?\s*)', '', result)
        result = re.sub(r"'", '', result) # Удалим одиночную кавычку.
        # Проверка наличия упоминания о компании в result
        if company is None or company == "None":
            print(f"{current_time()} Название компании не найдено. Удаляем файл.")
            os.remove(Link)
            print(f"{current_time()} Файл {Link} успешно удален.")
        else:
            # Проверка наличия упоминаний о компании и ключевых слов в result
            if company in result.lower():
                # Проверка наличия ключевых слов для счета
                if any(re.search(r'\b' + re.escape(keyword.lower()) + r'\b', result.lower()) for keyword in payment_invoice):
                    print(f"{current_time()} Это счёт на оплату.")
                    write_to_file(Документ="Счёт")
                    write_to_file(Link=Link)
                    write_to_file(result=result)
                    if re.search(r'числе ндс|ч\. ндс|ндс \(20|с ндс', result.lower()):
                        print(f"{current_time()} НДС есть.")
                        write_to_file(НДС="Есть")
                    elif re.search(r'без ндс', result.lower()):
                        print(f"{current_time()} Без НДС")
                        write_to_file(НДС="Нет")
                    else:
                        print(f"{current_time()} Не удалось определить наличие НДС в документе.")
                        write_to_file(НДС="Не_найдено.")
                # Проверка наличия ключевых слов для коммерческого предложения
                elif any(re.search(r'\b' + re.escape(keyword.lower()) + r'\b', result.lower()) for keyword in commercial_proposal):
                    print(f"{current_time()} Это коммерческое предложение.")
                    write_to_file(Документ="КП")
                    write_to_file(Link=Link)
                    write_to_file(result=result)
                    if re.search(r'числе ндс|ч\. ндс|ндс \(20|с ндс', result.lower()):
                        print(f"{current_time()} НДС есть.")
                        write_to_file(НДС="Есть")
                    elif re.search(r'без ндс', result.lower()):
                        print(f"{current_time()} Без НДС")
                        write_to_file(НДС="Нет")
                    else:
                        print(f"{current_time()} Не удалось определить наличие НДС в документе.")
                        write_to_file(НДС="Не_найдено.")
                else:
                    print(f"{current_time()} Нет упоминаний о КП или Счёте на оплату в документе. Удаляем файл.")
                    os.remove(Link)
            else:
                print(f"{current_time()} Название компании не найдено в документе. Удаляем файл.")
                os.remove(Link)
    except Exception as e:
        print(f"Ошибка В блоке process_and_delete_file(). Ошибка при обработке данных: {e}.")



def write_to_file(**data): # Функция для записи данных в файл.
    try:
        with open(file_path, 'a', encoding='utf-8') as file:
            for key, value in data.items():
                file.write(f"{key}: {value}\n")
            file.write("\n\n\n")  # Добавляем три пустые строки между записями
        print(f"{current_time()} Значение по ключу: {key}. Записано в файл: {file_path}")
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке write_to_file(). Ошибка при записи данных в файл: {e}.") 



def kup_get(driver,link): # Парсим что там на странице купца.
    try:
        # Открыть новую вкладку
        driver.execute_script("window.open('');")
        # Переключиться на новую вкладку
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(1)
        driver.get(link)
        print(f"{current_time()} Ссылка купца открыта.")
        time.sleep(3)
        # Проверка наличия элемента h1
        try:
            element = driver.find_element(By.TAG_NAME, "h1")
            # Проверка на заполнение заявки. Если она заполнена, то выходим.
            if "Спасибо за обратную связь." in element.text:
                print(f"{current_time()} Заявка уже заполнена.")
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(1)
                cup_exit(driver)
                main()
        except NoSuchElementException as e:
            print(f"{current_time()} Заявка еще не заполнена.")
        # Поиск элемента по XPath
        company = driver.find_element(By.XPATH, '//h3[contains(text(), "Вам поступил запрос на закупку товара от")]').text
        # Получение текста из элемента
        company = re.sub(r'Вам поступил запрос на закупку товара от ', '', company)
        cup_get(driver)
        driver.switch_to.window(driver.window_handles[0])
        write_to_file(Компания=company)
    except NoSuchElementException as e:
        print(f"{current_time()} Ссылка на купце просрочена, пометим письмо проблемным.")
        # Закрываем страницу с купцом.
        driver.close()
        # Сообщим браузеру что мы вернулись в главную вкладку.
        driver.switch_to.window(driver.window_handles[0])
        # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
        text = "Ссылка на купце просрочена."
        current_url = driver.current_url
        asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
        mark_an_error(driver)  # Пометим письмо проблемным. Нет упоминаний о компании или не счёт или кп.
        main()
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке kup_get(): {e}.")



def found_letter_is_correct(driver): # Проверяем что напарсили. Если нет упоминаний о компании или не счёт или кп. Помечаем письмо проблемным
    try:
        Link = read_from_file('Link')
        if not Link:
            # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
            text = "Нет упоминаний о компании или в прикреплённых файлах не указано что это Счёт на оплату или КП."
            current_url = driver.current_url
            asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
            # Вызвать функцию, если Link пуст значит не записали распознанное из прикреплённого письма
            print(f"{current_time()} Письмо не распознано. Пометим письмо проблемным.")
            mark_an_error(driver)  # Пометим письмо проблемным. Нет упоминаний о компании или не счёт или кп.
            print(f"{current_time()} Выходим из купца.")
            cup_exit(driver)
            main()
        # Если упоминание о компании есть обращаемся узнаем сколько товаров предлагает поставщик, и сколько в заявке купца.
        calculate_supplier_positions()
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке found_letter_is_correct(): {e}.")



def read_from_file(key): # Функция читает файл и возвращает значения по ключу.
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            value = []
            found_key = False
            empty_line_count = 0
            for line in lines:
                if found_key:
                    if line.strip():  # Если строка не пустая
                        value.append(line.strip())
                        empty_line_count = 0  # Сбрасываем счетчик пустых строк
                    else:
                        empty_line_count += 1  # Увеличиваем счетчик пустых строк
                        if empty_line_count >= 3:  # Если было уже 3 пустые строки, завершаем чтение
                            break
                elif line.startswith(f"{key}:"):
                    found_key = True
                    value.append(line.split(":", 1)[1].strip())
                    empty_line_count = 0  # Сбрасываем счётчик пустых строк после нахождения ключа
        return '\n'.join(value)
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке read_from_file(). Ошибка при чтении файла: {e}.")



def cup_exit(driver):  # Функция выходит из купца.
    try:
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(0.5)
        script = """
            var shadowRoot = document.querySelector("#ROOT-2521314 > vaadin-vertical-layout > vaadin-vertical-layout > vaadin-horizontal-layout:nth-child(1) > vaadin-vertical-layout > vaadin-menu-bar").shadowRoot;
            var button = shadowRoot.querySelector("div > vaadin-menu-bar-button:nth-child(1)");
            button.click();
        """
        driver.execute_script(script)
        time.sleep(0.5)
        exit_button = driver.find_element(By.XPATH, '//*[contains(text(), "Выйти")]')
        exit_button.click()
        time.sleep(0.5)
        logout_button = driver.find_element(By.ID, 'kc-logout')
        logout_button.click()
        time.sleep(0.5)
        driver.get("https://in.kupetscrm.ru/")
        time.sleep(2)
        while True:
            try:
                driver.find_element(By.XPATH, '//h1[contains(text(), "Войдите в свой аккаунт")]')
                break  # Если элемент найден, выход из цикла
            except NoSuchElementException:
                print(f"{current_time()} Выход из купца не успешен. Новая попытка.")
                time.sleep(1)
                driver.execute_script(script)
                time.sleep(0.5)
                exit_button = driver.find_element(By.XPATH, '//*[contains(text(), "Выйти")]')
                exit_button.click()
                time.sleep(0.5)
                logout_button = driver.find_element(By.ID, 'kc-logout')
                time.sleep(2)
                logout_button.click()
                time.sleep(2)
                back_link = driver.find_element(By.LINK_TEXT, '« Назад в приложение')
                # Клик по найденной ссылке
                back_link.click()
                time.sleep(2)
                driver.refresh()
        print(f"{current_time()} Выход из купца осуществлён. Вкладка закрыта.")
        driver.close() 
        driver.switch_to.window(driver.window_handles[0])
    except Exception as e:
        print(f"{current_time()} Ошибка В блоке cup_exit(): {str(e)}.")



def cup_get(driver): # Извлекаем и записываем в файл что требуется купцу. Кроме комментариев клиента, они нужны поставщику.
    try:
        combined_text = ""
        i = 0
        while True:
            selector = f'[slot="vaadin-grid-cell-content-{i}"]'
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            if elements:
                # Проверяем, не является ли элемент одним из элементов, которые нужно пропустить
                if i >= 25 and (i - 25) % 8 == 0:
                    i += 1
                    continue
                combined_text += ' '.join(element.get_attribute('textContent') for element in elements) + ' '
            else:
                break
            i += 1
        combined_text = re.sub(r'      ', '\n', combined_text)
        combined_text = re.sub(r'   ', '\n', combined_text)
        combined_text = re.sub(r'Нет в наличии', '\n', combined_text).strip() # Причёсываем текст от "Нет в наличии" и пробелов в начале и в конце
        combined_text = re.sub(r'шт', 'шт.\n', combined_text) # Причёсываем текст от "шт" переведем на новую строку
        combined_text = re.sub(r'in', 'шт.\n', combined_text) # Причёсываем текст от "in" переведем на новую строку (по другим вариантам)
        combined_text = combined_text.strip().replace('\n ', '\n') # Причёсываем текст от пробелов в начале и в конце и переведем на новую строку
        combined_text = re.sub(r'\n+', '\n', combined_text)
        combined_text = re.sub(r'^\s+', '', combined_text, flags=re.MULTILINE) # Удаление пробелов вначале строки
        combined_text = re.sub(r'^\.$', '', combined_text, flags=re.MULTILINE) # Удаление одиноко стоящей точки вначале строки.
        combined_text = re.sub(r'^\s*$', '', combined_text, flags=re.MULTILINE) # Удаление пустых строк.
        if combined_text:
            delivery = cup_get_delivery(driver) # Проверим нужна ли доставка клиенту.
            if delivery:
                combined_text = combined_text + '\n' + delivery
            # Вызываем функцию write_to_file с объединенным текстом
            write_to_file(Купец=combined_text)
        else:
            print(f"{current_time()} Не найдено элементов с текстом.")
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке cup_get(): {str(e)}.")



def cup_get_delivery(driver): # Проверим нужна ли доставка клиенту.
    try:
        delivery = driver.find_elements(By.XPATH, "//*[contains(text(), 'Клиент запросил доставку на адрес')]")
        
        # Вывод найденных элементов
        if delivery:
            return "Клиент запросил доставку до адреса."
        else:
            return None
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке cup_get_delivery(): {str(e)}.")




def send_request(database, topic): # Отправляем базу данных и тему вопроса нейросети. Она сравнивает товары и возвращает ответ.
    url = "http://192.168.1.16:5001/answer"  # URL сервера
    prompt = '''Ты занимаешься проверкой позиций на счетах на оплату. Твой ответ должен состоять образцов если это запрос о наличии товара твой
    ответ выглядит так [] В скобках товар в следующих скобках [] если аналог то укажи СЧЁТ ОШИБОЧЕН. Твой ответ обрабатывает скрипт Python отвечай
    на русском языке. Никаких дополнительных данных не указывай. Товар верен, цена. А так же в целом верен ли счёт. Перепроверь себя несколько раз.
    Надо точно знать что товары в счёте совпадают с товарами в запросе. Товар в счёте может немного отличаться в названии, расшифруй что это за
    позиция, и сравни с требуемой позицией из запроса. Но если товары разные обязательно укажи что СЧЁТ НЕВЕРЕН. В названии может быть аббревиатура,
    постарайся распознать и её. Так же если требуется 1 шт. цена и сумма могут быть одинаковые, учти это. Будь внимателен.'''
    data = {"database": database,"prompt": prompt, "topic": topic}
    while True:
        try:
            response = requests.post(url, json=data)
            if response.status_code == 200:
                answer = response.json()['answer']
                if len(answer) < 3 and not any(c.isalpha() for c in answer):
                    raise Exception("Длина ответа меньше 3 символов, а так же нет букв.")
                return answer
            else:
                return "Ошибка: Не удалось получить ответ от сервера"
        except Exception as e:
            print(f"{current_time()} Повторный запрос send_request(): {str(e)}.")



def read_and_parse_file(): # Спрашиваем у нейросети что мы распознали в счёте на оплату или кп. Перебираем товары по одному, так повышается качество ответа. Если всё верно вызывает функцию о контактах.
    try:
        kup_products = read_from_file("Купец")  # Узнаем, какие товары нужно найти
        kup_products = '\n'.join(kup_products.splitlines()[1:])  # Удаляем первую строку, там информация о Наименовании, Кол-ве, Ед. - нам это не нужно
        key = read_from_file("Документ")  # Узнаем, что это: счёт или КП
        database = (read_from_file("result"))  # Распознанный документ.
        gen_answer = '' # Создадим переменную для ответов нейросети.
        # Перебираем товары и спрашиваем у нейросети по одному
        for product in kup_products.splitlines():
            # Формируем запрос в соответствии с типом документа
            if key == "Счёт":
                request_promt = "Товар: "+ product + '''. Сравни товар из запроса с товарами в счёте на оплату цену и общую сумму, если не можешь найти
                цену за товар значит счёт неверен. Информация о доставке тут не нужна, сравнивай только товары. Если позиция совпадает в конце ответа
                в новой строке укажи [СЧЕТ ВЕРЕН] если не совпадают в новой строке напиши [СЧЁТ НЕВЕРЕН]. Твой ответ обрабатывает скрипт, отвечай без
                повторений и используй максимально короткий ответ. Пример ответа: Название товара:[ТЕЛЕФОН], [ПОЗИЦИЯ ВЕРНА], Цена:[00руб.], Сумма:[00руб.]'''
                print(f"{current_time()} Запрос к нейросети о совпадении товара и цене: {product}")
                answer = send_request(database, request_promt)
                # Удаляем пустые строки из строки answer
                answer = '\n'.join(line for line in answer.splitlines() if line.strip())
                print(f"{current_time()} Ответ нейросети: {answer}")
                if "счёт неверен" in answer.lower() or "[счёт неверен]" in answer.lower() or "счет неверен" in answer.lower() or "[счет неверен]" in answer.lower():
                    print(f"{current_time()} Неверный счёт. Пометим письмо проблемным.")
                    # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
                    text = (f"Счёт на оплату неверен. Нейросеть не распознала товар. {product}.")
                    current_url = driver.current_url
                    asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
                    mark_an_error(driver)  # Пометим письмо проблемным. Неверный счёт или кп.
                    print(f"{current_time()} Выходим из купца.")
                    cup_exit(driver)
                    main()
                elif not("счёт верен" in answer.lower() or "[счёт верен]" in answer.lower() or "счет верен" in answer.lower() or "[счет верен]" in answer.lower()):
                    print(f"{current_time()} Нет упоминаний о верности счёта, повторный вызов.")
                    read_and_parse_file()
                elif "цена: 00 руб." in answer.lower() or "цена: 00руб" in answer.lower() or "цена:00руб" in answer.lower() or "цена: (не" in answer.lower():
                    print(f"{current_time()} Нейросеть не распознала цену. Пометим письмо проблемным.")
                    # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
                    text = (f"Счёт на оплату неверен. Нейросеть не распознала цену на товар. {product}.")
                    current_url = driver.current_url
                    asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
                    print(f"{current_time()} Неверный счёт. Пометим письмо проблемным.")
                    mark_an_error(driver)  # Пометим письмо проблемным. Неверный счёт или кп.
                    print(f"{current_time()} Выходим из купца.")
                    cup_exit(driver)
                    main()
                else:
                    print(f"{current_time()} Товар в счёте верен, цена есть.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    gen_answer += f"{product} --- {answer}\n"
            elif key == "КП":
                request_promt = "Товар: "+ product + '''. Сравни товар из запроса с товарами в  коммерческом предложении укажи цену и общую сумму,
                если не можешь найти цену за товар значит счёт неверен. Информация о доставке тут не нужна, сравнивай только товары. Если позиция
                совпадает в конце ответа в новой строке укажи [СЧЕТ ВЕРЕН] если не совпадают в новой строке напиши [СЧЁТ НЕВЕРЕН]. Цену укажи один
                раз в конце строки после [ПОЗИЦИЯ ВЕРНА]. Твой ответ обрабатывает скрипт, отвечай только как в примере. Пример ответа: Название
                товара:[ТЕЛЕФОН], [ПОЗИЦИЯ ВЕРНА], Цена:[00руб.], Сумма:[00руб.]'''
                print(f"{current_time()} Запрос к нейросети о наличии и цене: {product}")
                answer = send_request(database, request_promt)
                # Удаляем пустые строки из строки answer
                answer = '\n'.join(line for line in answer.splitlines() if line.strip())
                print(f"{current_time()} Ответ нейросети: {answer}")
                if "кп неверен" in answer.lower() or "[кп неверен]" in answer.lower() or "кп неверен" in answer.lower() or "[кп неверен]" in answer.lower():
                    print(f"{current_time()} Неверный кп. Пометим письмо проблемным.")
                    # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
                    text = (f"Неверный КП. Нейросеть не распознала товар. {product}.")
                    current_url = driver.current_url
                    asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
                    mark_an_error(driver)  # Пометим письмо проблемным. Неверный счёт или кп.
                    print(f"{current_time()} Выходим из купца.")
                    cup_exit(driver)
                    main()
                elif not("кп верен" in answer.lower() or "[кп верен]" in answer.lower() or "кп верен" in answer.lower() or "[кп верен]" in answer.lower()):
                    print(f"{current_time()} Нет упоминаний о верности кп, повторный вызов.")
                    read_and_parse_file()
                elif "цена: 00 руб." in answer.lower() or "цена: 00руб" in answer.lower() or "цена:00руб" in answer.lower() or "цена: (не" in answer.lower():
                    print(f"{current_time()} Нейросеть не распознала цену. Пометим письмо проблемным.")
                    # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
                    text = (f"Неверный КП. Нейросеть не распознала цену на товар. {product}.")
                    current_url = driver.current_url
                    asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
                    mark_an_error(driver)  # Пометим письмо проблемным. Неверный счёт или кп.
                    print(f"{current_time()} Выходим из купца.")
                    cup_exit(driver)
                    main()
                else:
                    print(f"{current_time()} Товар в счёте верен, цена есть.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    gen_answer += f"{product} --- {answer}\n"
        gen_answer = re.sub(r"`", '', gen_answer) # Нейросеть любит вставить ` там где они не нужны. Исправим это.
        gen_answer = re.sub(r"'", '', gen_answer) # Нейросеть любит вставить ' там где они не нужны. Исправим это.
        write_to_file(Товар=gen_answer)  # Запишем в файл.
        get_contact_details() # Получаем контактные данные.
        # Временно. Позже удалить отсюда выход!
        print(f"{current_time()} Выходим из купца.")
        cup_exit(driver)
        product_key() # Выведем инфу по поставщику.
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке: read_and_parse_file(): {str(e)}.")



def get_contact_details(): # Спрашиваем у нейросети контактные данные. В данный момент в качестве базы она получает содержимое письма а так же информацию из счёта или кп.
    try:
        data = (read_from_file("email") + read_from_file("result"))
        # Получаем контактные данные.
        contact = '''. Сообщи для таблицы очень коротко и ёмко, имя контактного лица, почту, телефон и информацию о доставке, 
        если каких до данных нет, так и пиши Доставка: нет. Чем короче ответ тем лучше, не больше 100 символов.'''
        print(f"{current_time()} Запрос к нейросети о контактных данных и доставке.")
        contactless = send_request(data, contact)
        while len(contactless) < 3 and not any(c.isalpha() for c in contactless): # Иногда нейросеть криво отвечает, повторяем запрос.
            contactless = send_request(data, contact)
        # Удалим строки в которых нет полезной информации.
        if "телефон: не" in contactless.lower() or "телефон: нет" in contactless.lower():
            # Удаляем строку об отсутствии телефона.
            contactless = '\n'.join(line for line in contactless.splitlines() if "телефон:" not in line.lower())
        if "доставка: не" in contactless.lower() or "доставка: нет" in contactless.lower():
            # Удаляем строку об отсутствии доставки.
            contactless = '\n'.join(line for line in contactless.splitlines() if "доставка:" not in line.lower())
        if "лица: не" in contactless.lower() or "лицо: не" in contactless.lower():
            # Удаляем строку об отсутствии контакта.
            contactless = '\n'.join(line for line in contactless.splitlines() if "лицо:" not in line.lower())
        print(f"{current_time()} Ответ о контактных данных и доставке: {contactless}")
        write_to_file(Контакты_и_доставка=contactless)
        # Узнаем где находится товар.
        print(f"{current_time()} Запрос к нейросети о складах поставщика.")
        request = '''. Назови Город где находится склады поставщика, если у компании много Городов где есть склады, указывай
        все города через запятую, без собственных комментариев. Если непонятно где находится склад, указывай город поставщика.
        Если и город поставщика непонятен, дай ответ не известно.'''
        cargo = send_request(data, request)
        if "неизвестно" in cargo.lower() or "не знаю" in cargo.lower():
            print(f"{current_time()} Город товара неизвестен.")
        else:
            cargo = re.findall(r'[А-ЯЁ][а-яё]+', cargo) # Чистим вывод от лишнего ответа нейросети.
            cargo = list(set(cargo)) # Удалим одинаковые города из ответа.
            cargo = ', '.join(cargo)  # Добавляем запятые между городами
            cargo = re.sub(r'[\[\]\'\"]', '', cargo) # Очистим ответ от скобок.
            cargo += '.'  # Добавляем точку в конце строки
            print(f"{current_time()} Город где находится товар. {cargo}")
            write_to_file(Город=cargo)
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке: get_contact_details(): {str(e)}.")



def product_key(): # Сообщает по ключу название товара, цену и сумму.
    try:
        product = read_from_file("Товар")  # Узнаем, что за товары.
        key = read_from_file("Документ")  # Узнаем, что это: счёт или КП
        nds = read_from_file("НДС")  # Узнаем про НДС.
        kontakt = read_from_file("Контакты_и_доставка") # Узнаем контактную информацию.
        sity = read_from_file("Город") # Город склада или поставщика.
        if nds is None:
            nds = "НДС неизвестно."
        print(f"{current_time()} Это: {key}. НДС: {nds}.")
        lines = product.split('\n') # Разбиваем текст на строки
        index = 0  # Индекс текущей строки
        while index < len(lines):  # Пока не достигнут конец списка строк
            current_line = lines[index]  # Получаем текущую строку
            index += 1  # Увеличиваем индекс для перехода к следующей строке
            current_product = re.match(r'^(.*?)---', current_line, re.DOTALL).group(1).strip() # Взяли название до ---
            current_product = re.sub(r'\s+\d+\s*\S*$', '', current_product) # Удалим шт. и количество перед.
            current_product = re.sub(r'\b\d+\s*листов', '', current_product).rstrip() # Удалим Листов и цифры с количеством перед.
            price = re.sub(r'[\[\](){}]', '', current_line) # Очистим текст от скобок которые любит нейросеть.
            price = price.lower() # Сделаем текст маленькими буквами для простоты поиска.
            current_price = re.search(r'цена:(\d+)', price) # Вытаскиваем цену.
            if current_price:
                price_digits = current_price.group(1) # Чистим вывод цены от мусора.
            the_amount = re.search(r'сумма:(\d+)', price)
            if the_amount:
                amount_digits = the_amount.group(1) # Чистим вывод цены от мусора.
            print(f"{current_time()} Название товара: {current_product}. Цена за шт: {price_digits}. Общая сумма: {amount_digits}.")  # Выводим текущую строку (в данном примере)
        print(f"{current_time()} Контактные данные: {kontakt} Склад поставщика: {sity}")
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке: product(): {str(e)}.") 



def calculate_supplier_positions(): # Узнаем сколько товаров в счёте на оплату и в кп.
    try:
        text_result = (read_from_file("result"))
        key = read_from_file("Документ")  # Узнаем, что это: счёт или КП
        if key == "Счёт":
            pattern = r'^(?=.*(?:Товары? \(|Товар \())'
            matches = re.finditer(pattern, text_result, re.MULTILINE)
            for match in matches:
                line_number = text_result.count('\n', 0, match.start()) + 1
            lines = text_result.strip().split('\n')
            line_n = line_number
            text_result = '\n'.join(lines[line_n:])
            text_result = re.sub(r'Итого:.*[\s\S]*', '', text_result) # Удалим все строки начиная с Итого
            # Разделяем текст на строки
            lines = text_result.strip().split('\n')
            # Создаем список для хранения строк, которые останутся после фильтрации
            filtered_lines = []
            # Перебираем строки и проверяем, состоит ли каждая из них только из цифр
            for line in lines:
                # Проверяем, есть ли в строке символы, отличные от цифр и пробелов
                if not any(c.isalpha() or c == '.' for c in line):
                    continue  # Пропускаем строки, состоящие только из цифр и пробелов
                filtered_lines.append(line)
            # Объединяем отфильтрованные строки обратно в текст
            result = '\n'.join(filtered_lines)
            result = re.sub(r'^\D.*$\n?', '', result, flags=re.MULTILINE)
            num_lines = len(result.strip().split('\n'))
            kup_products = read_from_file("Купец")  # Узнаем, какие товары нужно найти
            kup_products = '\n'.join(kup_products.splitlines()[1:])  # Удаляем первую строку, там информация о Наименовании, Кол-ве, Ед. - нам это не нужно
            kup_lines = len(kup_products.split('\n'))
            if num_lines == kup_lines:
                # Временно, тут надо добавить проверку на доставку.
                print(f"{current_time()} Поставщик предлагает столько же позиций сколько необходимо покупателю. Позиции поставщика: {num_lines}. Купцу необходимо: {kup_lines}.")
                read_and_parse_file()
            elif num_lines > kup_lines:
                print(f"{current_time()} Количество позиций которые предлагает поставщик превосходят количество запросов купцу. Позиции поставщика: {num_lines}. Купцу необходимо: {kup_lines}. Пометим письмо.")
                # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
                text = (f"Счёт на оплату не верен. Позиции поставщика превышают количество запросов купцу. Позиции поставщика: {num_lines}. Купцу необходимо: {kup_lines}.")
                current_url = driver.current_url
                asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
                mark_an_error(driver)  # Пометим письмо проблемным. Нет упоминаний о компании или не счёт или кп.
                main()
            elif num_lines < kup_lines:
                print(f"{current_time()} Поставщик способен удовлетворить только часть запроса покупателя. Позиций у поставщика: {num_lines}. Купцу необходимо: {kup_lines}.")
                find_matching_products(kup_products, num_lines)
        elif key == "КП": 
            pass # Временно. Тут нужно найти и создать паттерны в КП для вычленения количества товаров в кп.



    except Exception as e:
        print(f"{current_time()} Ошибка в блоке: calculate_supplier_positions(): {str(e)}.") 



def find_matching_products(kup_products, num_lines): # Спрашиваем у нейросети по частичному совпадению мы распознали в счёте или кп.
    try:
        key = read_from_file("Документ")  # Узнаем, что это: счёт или КП
        database = (read_from_file("result"))  # Распознанный документ.
        good_answer = '' # Создадим переменную для положительных ответов нейросети.
        bad_answer = '' # Создадим переменную для отрицательных ответов нейросети.
        # Перебираем товары и спрашиваем у нейросети по одному
        for product in kup_products.splitlines():
            # Формируем запрос в соответствии с типом документа
            if key == "Счёт":
                request_promt = "Товар: "+ product + '''. Сравни товар из запроса с товарами в счёте на оплату цену и общую сумму, если не можешь найти
                цену за товар значит счёт неверен. Информация о доставке тут не нужна, сравнивай только товары. Если позиция совпадает в конце ответа
                в новой строке укажи [СЧЕТ ВЕРЕН] если не совпадают в новой строке напиши [СЧЁТ НЕВЕРЕН]. Твой ответ обрабатывает скрипт, отвечай без
                повторений и используй максимально короткий ответ. Пример ответа: Название товара:[ТЕЛЕФОН], [ПОЗИЦИЯ ВЕРНА], Цена:[00руб.], Сумма:[00руб.]'''
                print(f"{current_time()} Запрос к нейросети о совпадении товара и цене: {product}")
                answer = send_request(database, request_promt)
                # Удаляем пустые строки из строки answer
                answer = '\n'.join(line for line in answer.splitlines() if line.strip())
                print(f"{current_time()} Ответ нейросети: {answer}")
                if "счёт неверен" in answer.lower() or "[счёт неверен]" in answer.lower() or "счет неверен" in answer.lower() or "[счет неверен]" in answer.lower():
                    print(f"{current_time()} Товара в счёте нет.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    bad_answer += f"{product} --- {answer}\n"
                elif "цена: 00 руб." in answer.lower() or "цена: 00руб" in answer.lower() or "цена:00руб" in answer.lower() or "цена: (не" in answer.lower():
                    print(f"{current_time()} Цена товара в счёте отсутствует.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    bad_answer += f"{product} --- {answer}\n"
                elif "[]" in answer.lower() or "Позиция не найдена," in answer.lower() or "Цена:[]" in answer.lower() or "Сумма:[]" in answer.lower():
                    print(f"{current_time()} Нет упоминаний о товаре.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    bad_answer += f"{product} --- {answer}\n"
                elif "[СЧЕТ ВЕРЕН]" in answer.lower():
                    print(f"{current_time()} Товар в счёте верен.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    good_answer += f"{product} --- {answer}\n"
            elif key == "КП":
                request_promt = "Товар: "+ product + '''. Сравни товар из запроса с товарами в  коммерческом предложении укажи цену и общую сумму,
                если не можешь найти цену за товар значит счёт неверен. Информация о доставке тут не нужна, сравнивай только товары. Если позиция
                совпадает в конце ответа в новой строке укажи [СЧЕТ ВЕРЕН] если не совпадают в новой строке напиши [СЧЁТ НЕВЕРЕН]. Цену укажи один
                раз в конце строки после [ПОЗИЦИЯ ВЕРНА]. Твой ответ обрабатывает скрипт, отвечай только как в примере. Пример ответа: Название
                товара:[ТЕЛЕФОН], [ПОЗИЦИЯ ВЕРНА], Цена:[00руб.], Сумма:[00руб.]'''
                print(f"{current_time()} Запрос к нейросети о наличии и цене: {product}")
                answer = send_request(database, request_promt)
                # Удаляем пустые строки из строки answer
                answer = '\n'.join(line for line in answer.splitlines() if line.strip())
                print(f"{current_time()} Ответ нейросети: {answer}")
                if "счёт неверен" in answer.lower() or "[счёт неверен]" in answer.lower() or "счет неверен" in answer.lower() or "[счет неверен]" in answer.lower():
                    print(f"{current_time()} Товара в счёте нет.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    bad_answer += f"{product} --- {answer}\n"
                elif "цена: 00 руб." in answer.lower() or "цена: 00руб" in answer.lower() or "цена:00руб" in answer.lower() or "цена: (не" in answer.lower():
                    print(f"{current_time()} Цена товара в счёте отсутствует.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    bad_answer += f"{product} --- {answer}\n"
                elif "[]" in answer.lower() or "Позиция не найдена," in answer.lower() or "Цена:[]" in answer.lower() or "Сумма:[]" in answer.lower():
                    print(f"{current_time()} Нет упоминаний о товаре.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    bad_answer += f"{product} --- {answer}\n"
                elif "[СЧЕТ ВЕРЕН]" in answer.lower():
                    print(f"{current_time()} Товар в счёте верен.")
                    answer = answer.replace('\n', ' ') # Удаляем переносы строк.
                    good_answer += f"{product} --- {answer}\n"
        write_to_file(Товар_отсутствует=bad_answer)  # Запишем в файл отсутствующие позиции.
        good_answer = re.sub(r"`", '', good_answer) # Нейросеть любит вставить ` там где они не нужны. Исправим это.
        good_answer = re.sub(r"'", '', good_answer) # Нейросеть любит вставить ' там где они не нужны. Исправим это.
        good_answer_lines = len(good_answer.split('\n'))
        if good_answer_lines == num_lines:
            print(f"{current_time()} Все товары в счёте верны. Поставщик предлагает: {num_lines} совпадающих позиции: {good_answer_lines}.")
            write_to_file(Товар=good_answer)  # Запишем в файл.
            get_contact_details() # Получаем контактные данные.
            # Временно. Позже удалить отсюда выход!
            print(f"{current_time()} Выходим из купца.")
            cup_exit(driver)
            product_key() # Выведем инфу по поставщику.
        else:
            print(f"{current_time()} Не все позиции которые предлагает поставщик совпадают с товарами в счёте. Поставщик предлагает: {num_lines} совпадающих позиции: {good_answer_lines}. Пометим письмо.")
            # Отправим сообщение с ошибкой и ссылкой на письмо в Telegram.
            text = (f"Не все позиции которые предлагает поставщик совпадают с товарами в счёте. Поставщик предлагает: {num_lines} совпадающих позиции: {good_answer_lines}. Вот позиции которые совпадают с товарами в счёте: {good_answer}. Вот позиции которые не совпадают с товарами в счёте: {bad_answer}.")
            current_url = driver.current_url
            asyncio.run(send_message_to_user(token, chat_id, current_url, text)) 
            mark_an_error(driver)  # Пометим письмо проблемным. Неверный счёт или кп.
            print(f"{current_time()} Выходим из купца.")
            cup_exit(driver)
            main()
    except Exception as e:
        print(f"{current_time()} Ошибка в блоке: find_matching_products(): {str(e)}.")



try:
    # Переводим браузер на русский язык чтоб страницы открывались на русском. 
    add_russian_language(driver)
    # Проверяем авторизацию. Если нет, то авторизуемся.
    authorization(driver)
    if __name__ == "__main__": # Проверка имени модуля. Если имя модуля "__main__", то выполняется программа.
        main()



except Exception as ex: # Вывод исключения (если оно произошло)
    print(current_time(),"Вывод всех исключений:", ex)



finally: # Закрытие текущего окна браузера
    driver.close() # Закрытие текущего окна браузера в сессии WebDriver
    driver.quit() # Завершение сеанса WebDriver (выход из браузера) и освобождение ресурсов



"""
Опишу что происходит в коде. Перед основным запуском хром запускается на английском, в нижнем блоке try в самом конце программы происходит запуск двух функций.
add_russian_language(driver). Переводим браузер на русский язык чтоб страницы открывались на русском. В функции откроется страница почты, и в зависимости от
языка на странице, функция поймёт на русском мы или нет. Далее authorization(driver) Проверяем авторизацию. Если нет, то авторизуемся. К сожалению на mail.ru
подтянуть cookie отдельно не вышло, поэтому я сохранил все действия браузера. Создаётся папка в корне файла, где хранятся все настройки браузера и cookie в 
том числе. После того как отработают обе функции, вызывается main(). С этого момента цикл бесконечен. В main() заходит на саму почту. Первой вызывается 
функция filter(driver), она отфильтрует список почты, откроет только непрочитанные письма со вложениями, передаст управление в get_email_count(driver) эта 
функция игнорирует письма с красным флажком, парсит список писем, и отправляет список непрочитанных писем со вложениями и без флага в следующую функцию. Если
писем нет вернётся в main(). main() подождёт и повторит цикл с вызовом функции filter() и get_email_count(). Если письма в get_email_count() найдены список
будет обработан в open_last_investment_email(). В open_last_investment_email() произойдёт следующее так как ссылки на письма работают только 1 клик, второй
раз открыть новое письмо не выйдет, то я решил брать в работу самое старое письмо из новых. open_last_investment_email() откроет последнее письмо, очистит 
папку /root/Downloads/ от скачанных в прошлом файлов, они больше не нужны. Дальше она выдёргивает несколько параметров со страницы и с помощью функции
write_to_file() записывает их в файл. Она сохраняет называние входящей почты, получает со страницы всю переписку, что есть в письме. Но распознанный текст 
выглядит ужасно, перед записью текста в файл его причёсывает функция found_letter_is_correct(). Но перед записью новых данных в файл надо точно понимать есть
ли в этом смысл. Из функции open_last_investment_email() вызывается функция link_found(). Функция link_found() ищет на странице с письмом ссылку на купца. 
Если ссылки нет функция вызовет mark_an_error(). mark_an_error() пометит письмо проблемным, а именно сделает его непрочитанным чтоб на него обратил внимание
человек, и поставит красный флажок чтоб не пытаться его обработать снова. Если ссылка есть link_found() вернёт в функцию open_last_investment_email() ссылку
на купца. Та узнав что ссылка есть, запишет доступную информацию со страницы и передаст ссылку на купца в функцию kup_get(). kup_get() считает название 
компании в будущем нам название очень понадобится. Так же kup_get() вызовет kup_get_product() эта функция считает список товаров и запишет это в файл. Итого
что мы имеем на данный момент, в файле /root/Downloads/output.txt которую обрабатывает функция write_to_file() мы имеем список требуемых купцу товаров под 
ключом (Купец), название компании под ключом (Компания), а так же почту под ключом (email) и переписку под ключом (Содержимое_письма). Тут важно после того
как мы всё это узнали и записали в файл у нас осталась открыта вторая вкладка в фоне с купцом и ссылка в переменной на вложение и мы снова в 
open_last_investment_email() важно именно отсюда дальше запускать работу со вложением вызывается  функция link_search(), она ищет ссылку на вложенные файлы,
мы точно знаем что ссылка есть, если ссылку не находит link_search() программа останавливается для отлова ошибки, и ссылка на нераспознанное письмо
записывается в файл под ключом (link_search), так как во вложении могут быть несколько файлов, мы находим ссылку сразу на все файлы. Если всё хорошо и ссылка
найдена функция link_search() возвращает ссылку в open_last_investment_email(), оттуда вызов к скачиванию. Функция open_last_investment_email() скидывает 
ссылку на архив или вложение если файл один в функцию open_link_in_browser(). open_link_in_browser() скачивает файл на всякий случай убирает из названий 
пробелы и ставит _ и отправляет список переименованных файлов дальше в process_downloaded_file(). Тут происходит анализ скачанного файла в зависимости 
от расширения, process_pdf() переводит pdf в текст, process_image() переводит текст с картинок в текст, мало ли счёт в фото. ".jpg", ".jpeg", ".png", ".gif".
Следующая функция read_excel() так же переводит ".xlsx", ".xlsm", ".xlsb", ".xls" в текст и последняя функция с которой работает process_downloaded_file(),
это unzip_archive() если поведение функций с распознаванием понятна, то unzip_archive() делает следующее разбирает и удаляет архив оставляя только файлы из
архива, вызывает функцию rename_files() убирает из названий пробелы и ставит _ и отправляет список переименованных файлов дальше в process_files(), та по 
одному отправляет их в process_downloaded_file() там уже не архив и распознаются все файлы, но ещё не записывается результат. Перед записью результатов в файл
надо понять что мы вообще распознали, функции по распознаванию отправляют результаты в функцию process_and_delete_file() там происходит следующее, проверяется
есть ли упоминания о компании в распознанном документе. Если есть проверяем счёт это или кп записываем в ключ (Документ), что это счёт или кп, если документ 
не имеет отношения к счёту или кп нам не интересно письмо, но сразу пометить письмо нельзя вдруг документов несколько, а мы распознали только первый файл. 
Но сам файл на не нужен поэтому удаляем файл и забываем ссылку на него. Если всё же во вложении счёт или кп и компания нужная там фигурирует всё, записывается
в файл под ключом (result), а так же информацию с НДС или без под ключом (НДС), и ссылку на файл под ключом (Link) после того как распознаны и записаны все 
документы управление возвращается в open_last_investment_email() тут надо понимать что сюда мы вернулись после того как архив был скачан и разобран на текст
и если всё хорошо распознаны все файлы из архива записаны в файл два ключа (Link) (result), вызываем функцию found_letter_is_correct() она проверит есть ли в
файле /root/Downloads/output.txt что-то под ключом (Link), значит это счёт или кп и выставлен он на компанию из ключа (Компания). Если ключа (Link), значит мы
всё что мы распознали лажа вызовет функцию mark_an_error() та пометит письмо и вернётся в found_letter_is_correct() там будет вызвана вторая функция cup_exit()
которая откроет вторую вкладку, выйдет из купца и закроет вкладку, после вернёт управление в found_letter_is_correct() и последней будет вызвана main() для 
нового цикла. А что если found_letter_is_correct() найдёт ключ (Link) в файле /root/Downloads/output.txt тогда вызовем read_and_parse_file() которая будет 
спрашивать у нейросети что мы распознали, функция read_and_parse_file() переберёт каждый товар требуемый компании со страницы купца записанные в файле под 
ключом (Купец), она методом перебора спросит о каждом товаре отдельно, так как спрашивать списком сразу нельзя, это повышает шансы на ошибку нейросети. 
Функция создаёт роль для нейросети базу для товаров база это счёт на оплату и запрос с товаром. Если этого не делать нейросеть галлюцинирует. Запросы 
обрабатываются удалённо через функцию send_request(). Сама же функция read_and_parse_file() отправляет запрос и проверяет ответ на несколько условий. Если 
нейросеть в ответе пишет что такого товара в счёте или кп нет, помечает и вызывает функции выхода из купца и пометку письма проблемным. Или если нейросеть
ответила что товар такой есть, но не смогла сказать сколько он стоит так же выходим из купца и помечаем письмо проблемным. Если нейросеть написала что товар
верен и указала цену из счёта или кп, записываем это в файл под ключом (Товар), запишется название из купца и цена которую указала нейросеть. Далее вызов
функции get_contact_details(). Функция get_contact_details() возьмёт несколько ключей, почту, текст из письма и текст из счёта или кп соберёт это в базу 
и отправит запрос о контактный информации телефон, почта, контактное лицо, информация о доставке. Ответ запишет в файл под ключом (Контакты_и_доставка). 
Запросы по ключам обрабатывает функция read_from_file() она возвращает значения по ключу из файла. Важно понимать что функция read_and_parse_file() 
отработает нормально только и передаст управление дальше в функции ввода цен и файлов на странице купца только в том случае, если она верно распознала все
товары и цены из списка купца. Если нейросеть не поняла хотя бы один товар или цену товара, письмо будет помечено проблемным. Так же реализована функция 
которая получает количество строк с товарами купца и сравнивает с количеством строк в файле на оплату. И если товаров в счёте или кп больше чем товаров 
в купце письмо будет помечено проблемным, и выход из купца. Если в счёте или кп меньше чем в купце письмо будет обработано, произойдёт сверка каждой позиции
поставщика со списком купца. Если все позиции поставщика присутствуют в запросе клиента письмо будет обработано.
Осталось несколько финальных функций которые впишут цены комментарии контактную информацию и прикрепят файл и отправят закупцу.
"""